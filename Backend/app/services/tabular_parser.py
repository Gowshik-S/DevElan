import csv
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook


ALLOWED_TABULAR_EXTENSIONS = {".csv", ".xlsx"}
REGISTER_NO_PATTERN = re.compile(r"^[0-9]{2}[A-Z]{3}[0-9]{3}$", flags=re.IGNORECASE)
USE_CASE_CODE_PATTERN = re.compile(r"^EL-\d+$", flags=re.IGNORECASE)
KNOWN_HEADER_TOKENS = {
    "id",
    "name",
    "full_name",
    "student_name",
    "register",
    "register_no",
    "register_number",
    "reg_no",
    "email",
    "email_id",
    "mail",
    "class",
    "class_name",
    "assign_class",
    "year",
    "semester",
    "year_sem",
    "year_semester",
    "year_semester_value",
    "use_case_code",
    "usecase_code",
    "problem_statement_number",
    "problem_statement_no",
    "problem_statement_id",
    "problem_id",
    "problem_no",
    "problem_number",
    "problem_code",
    "topic",
    "title",
    "objective",
    "the_objective",
    "objective_text",
    "description",
    "problem_title",
    "problem_statement_title",
    "problem",
    "problem_description",
    "key_concepts",
    "key_concept",
    "workflow_steps",
    "workflow",
    "steps",
    "output",
    "expected_output",
    "output_description",
}


def _normalize_header(header: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", header.strip().lower())
    return cleaned.strip("_")


def _looks_like_header_row(cells: list[Any]) -> bool:
    normalized = [_normalize_header(str(cell or "")) for cell in cells]
    non_empty = [token for token in normalized if token]
    if not non_empty:
        return False

    if any(token in KNOWN_HEADER_TOKENS for token in non_empty):
        return True

    for cell in cells:
        text = str(cell or "").strip()
        if not text:
            continue
        if "@" in text:
            return False
        if REGISTER_NO_PATTERN.match(text.upper()):
            return False
        if USE_CASE_CODE_PATTERN.match(text.upper()):
            return False

    return True


def _rows_to_dicts(raw_rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not raw_rows:
        return []

    has_header = _looks_like_header_row(raw_rows[0])
    if has_header:
        headers = [
            _normalize_header(str(value or "")) or f"column_{index + 1}"
            for index, value in enumerate(raw_rows[0])
        ]
        data_rows = raw_rows[1:]
        starting_row = 2
    else:
        max_columns = max(len(row) for row in raw_rows)
        headers = [f"column_{index + 1}" for index in range(max_columns)]
        data_rows = raw_rows
        starting_row = 1

    rows: list[dict[str, Any]] = []
    for row_number, row_values in enumerate(data_rows, start=starting_row):
        normalized: dict[str, Any] = {"__row_number__": row_number}
        for index, header in enumerate(headers):
            value = row_values[index] if index < len(row_values) else None
            normalized[header] = value.strip() if isinstance(value, str) else value

        if any(
            value not in (None, "")
            for key, value in normalized.items()
            if key != "__row_number__"
        ):
            rows.append(normalized)

    return rows


def _read_csv(data: bytes) -> list[dict[str, Any]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV file must be UTF-8 encoded.",
        ) from exc

    reader = csv.reader(StringIO(text))
    raw_rows = [row for row in reader]
    if not raw_rows:
        return []

    return _rows_to_dicts(raw_rows)


def _read_xlsx(data: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.active

    raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    if not raw_rows:
        workbook.close()
        return []

    rows = _rows_to_dicts(raw_rows)
    workbook.close()
    return rows


def parse_uploaded_table(file: UploadFile) -> list[dict[str, Any]]:
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file must include a filename.",
        )

    extension = Path(file.filename).suffix.lower()
    if extension not in ALLOWED_TABULAR_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .csv and .xlsx files are supported.",
        )

    data = file.file.read()
    if not data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty.",
        )

    if extension == ".csv":
        return _read_csv(data)
    return _read_xlsx(data)
