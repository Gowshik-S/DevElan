from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook
from sqlalchemy import select

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.db.session import SessionLocal
from app.models.usecase import UseCase

USE_CASE_PATTERN = re.compile(r"^(EL-\d+)\.\s*(.+)$", flags=re.IGNORECASE)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_docx_usecases(file_path: Path) -> list[dict[str, Any]]:
    document = Document(str(file_path))
    lines = [_clean(paragraph.text) for paragraph in document.paragraphs]
    lines = [line for line in lines if line]

    records: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    reading_key_concepts = False

    for line in lines:
        match = USE_CASE_PATTERN.match(line)
        if match:
            if current is not None:
                records.append(current)
            current = {
                "code": match.group(1).upper(),
                "title": match.group(2).strip(),
                "description": "",
                "key_concepts": [],
                "workflow_steps": [],
                "output_description": None,
            }
            reading_key_concepts = False
            continue

        if current is None:
            continue

        lower = line.lower()
        if lower.startswith("problem statement"):
            content = re.sub(r"^problem\s*statement\s*", "", line, flags=re.IGNORECASE)
            current["description"] = content.strip(" :-\u2013")
            reading_key_concepts = False
            continue

        if lower.startswith("key concepts"):
            reading_key_concepts = True
            continue

        if lower.startswith("output"):
            content = re.sub(r"^output\s*", "", line, flags=re.IGNORECASE)
            current["output_description"] = content.strip(" :-\u2013\"")
            reading_key_concepts = False
            continue

        if reading_key_concepts:
            normalized = line.lstrip("-• ").strip()
            if normalized:
                current["key_concepts"].append(normalized)

    if current is not None:
        records.append(current)

    return records


def _normalize_headers(header_row: list[Any]) -> list[str]:
    return [re.sub(r"[^a-z0-9]+", "_", _clean(value).lower()).strip("_") for value in header_row]


def parse_tabular_usecases(file_path: Path) -> list[dict[str, Any]]:
    extension = file_path.suffix.lower()
    rows: list[dict[str, Any]] = []

    if extension == ".csv":
        with file_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
            reader = csv.DictReader(csv_file)
            for row in reader:
                normalized = {
                    re.sub(r"[^a-z0-9]+", "_", _clean(key).lower()).strip("_"): _clean(value)
                    for key, value in row.items()
                    if key
                }
                if any(normalized.values()):
                    rows.append(normalized)
    elif extension == ".xlsx":
        workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        worksheet = workbook.active
        values = [list(row) for row in worksheet.iter_rows(values_only=True)]
        workbook.close()
        if not values:
            return []

        headers = _normalize_headers(values[0])
        for value_row in values[1:]:
            normalized: dict[str, str] = {}
            for index, value in enumerate(value_row):
                header = headers[index] if index < len(headers) else f"column_{index + 1}"
                normalized[header] = _clean(value)
            if any(normalized.values()):
                rows.append(normalized)
    else:
        raise ValueError("Only .docx, .csv, and .xlsx files are supported.")

    records: list[dict[str, Any]] = []
    for row in rows:
        code = _clean(row.get("code") or row.get("use_case_code") or row.get("problem_code"))
        title = _clean(row.get("title") or row.get("use_case_title") or row.get("problem_title"))
        description = _clean(row.get("description") or row.get("problem_statement") or row.get("objective"))
        if not code:
            continue

        key_concepts_raw = _clean(row.get("key_concepts") or row.get("key_concept"))
        workflow_raw = _clean(row.get("workflow_steps") or row.get("workflow") or row.get("steps"))
        output_description = _clean(row.get("output_description") or row.get("output")) or None

        key_concepts = [item.strip() for item in key_concepts_raw.replace("\n", ";").replace(",", ";").split(";") if item.strip()]
        workflow_steps = [item.strip() for item in workflow_raw.replace("\n", ";").replace(",", ";").split(";") if item.strip()]

        records.append(
            {
                "code": code.upper(),
                "title": title or f"Use Case {code.upper()}",
                "description": description or "Imported via tabular file.",
                "key_concepts": key_concepts,
                "workflow_steps": workflow_steps,
                "output_description": output_description,
            }
        )

    return records


def parse_usecases(file_path: Path) -> list[dict[str, Any]]:
    extension = file_path.suffix.lower()
    if extension == ".docx":
        return parse_docx_usecases(file_path)
    if extension in {".csv", ".xlsx"}:
        return parse_tabular_usecases(file_path)
    raise ValueError("Unsupported file type. Use .docx, .csv, or .xlsx.")


def upsert_usecases(records: list[dict[str, Any]], overwrite: bool = True) -> tuple[int, int, int]:
    created_count = 0
    updated_count = 0
    skipped_count = 0

    with SessionLocal() as db:
        for record in records:
            code = _clean(record.get("code")).upper()
            if not code:
                skipped_count += 1
                continue

            existing = db.scalar(select(UseCase).where(UseCase.code.ilike(code)))
            title = _clean(record.get("title")) or f"Use Case {code}"
            description = _clean(record.get("description")) or "Imported via script."
            key_concepts = record.get("key_concepts") if isinstance(record.get("key_concepts"), list) else []
            workflow_steps = record.get("workflow_steps") if isinstance(record.get("workflow_steps"), list) else []
            output_description = _clean(record.get("output_description")) or None

            if existing is None:
                db.add(
                    UseCase(
                        code=code,
                        title=title,
                        description=description,
                        key_concepts=key_concepts,
                        workflow_steps=workflow_steps,
                        output_description=output_description,
                    )
                )
                created_count += 1
                continue

            if not overwrite:
                skipped_count += 1
                continue

            changed = False
            if existing.title != title:
                existing.title = title
                changed = True
            if existing.description != description:
                existing.description = description
                changed = True
            if key_concepts and existing.key_concepts != key_concepts:
                existing.key_concepts = key_concepts
                changed = True
            if workflow_steps and existing.workflow_steps != workflow_steps:
                existing.workflow_steps = workflow_steps
                changed = True
            if output_description is not None and existing.output_description != output_description:
                existing.output_description = output_description
                changed = True

            if changed:
                db.add(existing)
                updated_count += 1
            else:
                skipped_count += 1

        db.commit()

    return created_count, updated_count, skipped_count


def main() -> None:
    parser = argparse.ArgumentParser(description="Import and upsert use cases into DevEla database.")
    parser.add_argument("input_file", help="Path to .docx/.csv/.xlsx use-case file")
    parser.add_argument(
        "--no-overwrite",
        action="store_true",
        help="Do not update existing use cases; only create missing codes.",
    )
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    records = parse_usecases(input_path)
    if not records:
        raise SystemExit("No use case records were found in the input file.")

    created, updated, skipped = upsert_usecases(records, overwrite=not args.no_overwrite)
    print(
        f"usecases_processed={len(records)} created={created} updated={updated} skipped={skipped}"
    )


if __name__ == "__main__":
    main()
